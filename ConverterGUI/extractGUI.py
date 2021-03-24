import pandas as pd
import numpy as np
from openpyxl import load_workbook
import re
import datetime
import time
import numpy as np
# #list of word to check with regex for OS
# listOS=["IBM","Windows", "OPen SUSE","VMWare","Linux","Solaris","Red Hat","Centos","RedHat",
# "AIX","Z/OS","Win","RHEL","v5r4","v6r1", "OS400","OS390","OS/400","V7R","ZOS","AS400","V5R3","V7 R1"]
# #list of dict that will trigger if regex matching
# dictOS={"RHEL":"Linux","Red Hat":"Linux","RedHat":"Linux","AIX":"IBM","Z/OS":"IBM","ZOS":"IBM","V5R3":"IBM", "V7 R1": "IBM",
# "Win":"Windows","Centos":"Linux","v5r4":"IBM","v6r1":"IBM","OS400":"IBM","OS390":"IBM","OS/400":"IBM","V7R":"IBM","AS400":"IBM"}
# #path for file
# PATH="Newest.xlsx"
# #dic for critcality
# criticality_dict = {"Critiacl": 'Critical', "Crifical": 'Critical',"Very Important (3rd party system)":"Very Important"}
# ipo_dict = {"purchased": 'P',"outsourced": 'O', "inhouse": 'I',0:"O","O P":"O,P","Required":"P","BNM system":"BNM System"}
    

def formatExcel(path,dictFilter,sheetName,colName):
    #read excel sheet 
    data = pd.read_excel(path,sheet_name=sheetName)
    data = data.dropna(axis='columns',how="all")
    book = load_workbook(path)
    newPath=path.replace(".xlsx","")
    writer = pd.ExcelWriter("{}_backup.xlsx".format(newPath), engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data.to_excel(writer, sheetName, index=False)
    writer.save()
    writer.close()

    # print(list(data))
    data[colName],updateRecord=iterateReplaceWord(data,dictFilter,colName,path)
    #process to write excel
    book = load_workbook(path)
   
    writer = pd.ExcelWriter(path, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data.to_excel(writer, sheetName, index=False)
    writer.save()

    with open("{0} log.txt".format(newPath), "a") as txt_file:
        for i in range(len(updateRecord)):
            if i ==0:
                txt_file.write("\n")
            txt_file.write(str(updateRecord[i]) + "\n") # works with any number of elements in a line





def iterateReplaceWord(data,dictFilter,colName,path):
    updatedData=[]
    updateRecord=[]
    # current_time=datetime.datetime.now()
    # time.strftime('%l:%M%p %Z on %b %d, %Y')
    updateRecord.append(    time.ctime() )# 'Mon Oct 18 13:35:29 2010'

    updateRecord.append(path)

    print("dict: ",dictFilter)

    #iterate row by row using itertuples
    for i,row in data.iterrows(): 
        if data.at[i,colName] != data.at[i,colName]:
            if dictFilter.get("<Blank>"):
                updatedData.append(dictFilter.get("<Blank>"))
                strReport="ID Number:{0}: Updated Previous Value: {1}, Current Value: {2}, Column Name: {3}".format(
                    str(i+1),str("<Blank>"),dictFilter.get("<Blank>"),colName)
                updateRecord.append(strReport)
        elif dictFilter.get(str(data.at[i,colName])):
            updatedData.append(dictFilter.get(str(data.at[i,colName])))
            strReport="ID Number:{0}: Updated Previous Value: {1}, Current Value: {2}, Column Name: {3}".format(
                str(i+1),str(data.at[i,colName]),dictFilter.get(data.at[i,colName]),colName)
            updateRecord.append(strReport)
        else:
            updatedData.append(str(data.at[i,colName]))
    # print(value)
    return updatedData,updateRecord

def return_sheet(path):

    xl = pd.ExcelFile(path)
    return xl.sheet_names 

def return_column(path,sheet):
    xl = pd.ExcelFile(path)
    for i in xl.sheet_names :
        if i == sheet:
            data = pd.read_excel(path,sheet_name=i)
            data = data.dropna(axis='columns',how="all")


            return list(data)
    return


def return_value(path,sheet,col):
    value=set()
    data = pd.read_excel(path,sheet_name=sheet)
    data = data.dropna(axis='columns',how="all")
    for i,row in data.iterrows(): 
        value.add(data.at[i,col])
    # print(value)
    return value

def createDict(path,value,wordArray,col,sheet):
    dictFilter={}
    
    for i in wordArray:
        if i != i:

            print("yes")
            newDict={"<Blank>":value}
            dictFilter.update(newDict)

        else:
            print(type(i))
            if type(i)==np.int64:
                newDict={i:int(value)}
                dictFilter.update(newDict)
            else:
                newDict={str(i):value}
                dictFilter.update(newDict)
    print("Col",col)
    formatExcel(path,dictFilter, sheet,col)



